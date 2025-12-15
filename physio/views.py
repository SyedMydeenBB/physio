from django.shortcuts import render

def index(request):
    """Home page view"""
    context = {
        'title': 'CHENNAI PHYSIO CARE - For a Better Quality of Life',
        'phone': '+91-73052-74514',
        'email': 'admin@chennaiphysiocare.com',
        'address': '2, Pearl Apartments, 13th Main Rd, Anna Nagar West, Chennai, Tamil Nadu 600040',
        'years_experience': '20+',
        'patients_served': '25K+',
        'recovery_rate': '99.9%',
    }
    return render(request, 'physio/index.html', context)

def services(request):
    """Services page view"""
    context = {
        'title': 'Our Services - Chennai Physio Care',
    }
    return render(request, 'physio/services.html', context)

def contact(request):
    """Contact page view"""
    context = {
        'title': 'Contact Us - Chennai Physio Care',
    }
    return render(request, 'physio/contact.html', context)

def blog(request):
    """Blog page view"""
    context = {
        'title': 'Blog - Chennai Physio Care',
    }
    return render(request, 'physio/blog.html', context)

def guide(request):
    """Guide page view"""
    context = {
        'title': 'Patient Guide - Chennai Physio Care',
    }
    return render(request, 'physio/guide.html', context)

def about(request):
    """About page view"""
    context = {
        'title': 'About Us - Chennai Physio Care',
    }
    return render(request, 'physio/about.html', context)



from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Patient
from .forms import PatientForm, UploadFileForm
import pandas as pd
from io import BytesIO

def patient_list(request):
    patients = Patient.objects.all()
    return render(request, 'patient_list.html', {'patients': patients})

def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('patient_list')
    else:
        form = PatientForm()
    return render(request, 'patient_form.html', {'form': form})

def patient_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('patient_list')
    else:
        form = PatientForm(instance=patient)
    return render(request, 'patient_form.html', {'form': form})

def patient_delete(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        return redirect('patient_list')
    return render(request, 'patient_confirm_delete.html', {'patient': patient})

def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            df = pd.read_excel(file)

            # Normalize headers
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

            for _, row in df.iterrows():
                Patient.objects.update_or_create(
                    case_number=row.get('case_number'),
                    defaults={
                        'name': row.get('name'),
                        'age': row.get('age'),
                        'gender': row.get('gender'),
                        'chief_complaint': row.get('chief_complaint'),
                        'reference': row.get('reference'),
                        'contact': row.get('contact_number') or row.get('contact'),
                        'address': row.get('address')
                    }
                )
            return redirect('patient_list')
    else:
        form = UploadFileForm()
    return render(request, 'upload_excel.html', {'form': form})


def download_excel(request):
    patients = Patient.objects.all()
    data = {
        'Name': [p.name for p in patients],
        'Case Number': [p.case_number for p in patients],
        'Age': [p.age for p in patients],
        'Gender': [p.get_gender_display() for p in patients],
        'Chief Complaint': [p.chief_complaint for p in patients],
        'Reference': [p.reference for p in patients],
        'Contact': [p.contact for p in patients],
        'Address': [p.address for p in patients],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=patients.xlsx'
    return response

def dashboard(request):
    return render(request, 'dashboard.html')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import pandas as pd
from .models import DailySheet
from .forms import DailySheetForm, ExcelUploadForm

def daily_sheet_list(request):
    """List all daily sheets with filters"""
    sheets = DailySheet.objects.all()
    
    # Get filter parameters
    filter_type = request.GET.get('filter_type', 'all')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    search = request.GET.get('search', '')
    
    # Apply search filter
    if search:
        sheets = sheets.filter(
            Q(name__icontains=search) |
            Q(case_number__icontains=search) |
            Q(diagnosis__icontains=search)
        )
    
    # Apply date filters
    today = datetime.now().date()
    
    if filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        sheets = sheets.filter(date__range=[start_date, end_date])
    elif filter_type == 'month':
        sheets = sheets.filter(date__year=today.year, date__month=today.month)
    elif filter_type == 'year':
        sheets = sheets.filter(date__year=today.year)
    elif filter_type == 'custom' and custom_start and custom_end:
        sheets = sheets.filter(date__range=[custom_start, custom_end])
    
    context = {
        'sheets': sheets,
        'filter_type': filter_type,
        'custom_start': custom_start,
        'custom_end': custom_end,
        'search': search,
    }
    return render(request, 'daily_sheet_list.html', context)

def daily_sheet_create(request):
    """Create a new daily sheet entry"""
    if request.method == 'POST':
        form = DailySheetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Daily sheet entry created successfully!')
            return redirect('daily_sheet_list')
    else:
        form = DailySheetForm()
    
    return render(request, 'daily_sheet_form.html', {'form': form, 'action': 'Create'})

def daily_sheet_update(request, pk):
    """Update an existing daily sheet entry"""
    sheet = get_object_or_404(DailySheet, pk=pk)
    
    if request.method == 'POST':
        form = DailySheetForm(request.POST, instance=sheet)
        if form.is_valid():
            form.save()
            messages.success(request, 'Daily sheet entry updated successfully!')
            return redirect('daily_sheet_list')
    else:
        form = DailySheetForm(instance=sheet)
    
    return render(request, 'daily_sheet_form.html', {'form': form, 'action': 'Update', 'sheet': sheet})

def daily_sheet_delete(request, pk):
    """Delete a daily sheet entry"""
    sheet = get_object_or_404(DailySheet, pk=pk)
    
    if request.method == 'POST':
        sheet.delete()
        messages.success(request, 'Daily sheet entry deleted successfully!')
        return redirect('daily_sheet_list')
    
    return render(request, 'daily_sheet_confirm_delete.html', {'sheet': sheet})

def daily_sheet_export(request):
    """Export daily sheets to Excel"""
    # Get filtered data
    sheets = DailySheet.objects.all()
    
    filter_type = request.GET.get('filter_type', 'all')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    
    today = datetime.now().date()
    
    if filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        sheets = sheets.filter(date__range=[start_date, end_date])
    elif filter_type == 'month':
        sheets = sheets.filter(date__year=today.year, date__month=today.month)
    elif filter_type == 'year':
        sheets = sheets.filter(date__year=today.year)
    elif filter_type == 'custom' and custom_start and custom_end:
        sheets = sheets.filter(date__range=[custom_start, custom_end])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Sheets'
    
    # Define headers
    headers = [
        'Date', 'Name', 'Case Number', 'Diagnosis', 'Charge', 'Received',
        'Payment Status', 'Payment Type', 'Payment Frequency', 'In Time',
        'Out Time', 'Treatment 1', 'Treatment 2', 'Treatment 3', 'Treatment 4',
        'Therapist 1', 'Therapist 2'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_num, sheet in enumerate(sheets, 2):
        ws.cell(row=row_num, column=1).value = sheet.date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = sheet.name
        ws.cell(row=row_num, column=3).value = sheet.case_number
        ws.cell(row=row_num, column=4).value = sheet.diagnosis
        ws.cell(row=row_num, column=5).value = float(sheet.charge)
        ws.cell(row=row_num, column=6).value = float(sheet.received)
        ws.cell(row=row_num, column=7).value = sheet.get_payment_status_display()
        ws.cell(row=row_num, column=8).value = sheet.get_payment_type_display()
        ws.cell(row=row_num, column=9).value = sheet.get_payment_frequency_display()
        ws.cell(row=row_num, column=10).value = sheet.in_time.strftime('%H:%M') if sheet.in_time else ''
        ws.cell(row=row_num, column=11).value = sheet.out_time.strftime('%H:%M') if sheet.out_time else ''
        ws.cell(row=row_num, column=12).value = sheet.treatment_1
        ws.cell(row=row_num, column=13).value = sheet.treatment_2
        ws.cell(row=row_num, column=14).value = sheet.treatment_3
        ws.cell(row=row_num, column=15).value = sheet.treatment_4
        ws.cell(row=row_num, column=16).value = sheet.get_therapist_1_display()
        ws.cell(row=row_num, column=17).value = sheet.get_therapist_2_display() if sheet.therapist_2 else ''
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=daily_sheets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

def daily_sheet_import(request):
    """Import daily sheets from Excel"""
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                # Map payment status
                payment_status_map = {
                    'Paid': 'paid',
                    'Partially Paid': 'partially_paid',
                    'Not Paid': 'not_paid'
                }
                
                payment_type_map = {
                    'QR': 'qr',
                    'Cash': 'cash'
                }
                
                payment_frequency_map = {
                    'Daily Basis': 'daily',
                    'Weekly Basis': 'weekly',
                    'Monthly Basis': 'monthly'
                }
                
                therapist_map = {
                    'Dr. Basidh': 'dr_basidh',
                    'Dr. Visitra': 'dr_visitra',
                    'Dr. Bharatiselvi': 'dr_bharatiselvi'
                }
                
                imported_count = 0
                
                for index, row in df.iterrows():
                    try:
                        sheet = DailySheet(
                            date=pd.to_datetime(row['Date'], dayfirst=True).date(),
                            name=row['Name'],
                            case_number=str(row['Case Number']),
                            diagnosis=row['Diagnosis'],
                            charge=float(row['Charge']),
                            received=float(row['Received']),
                            payment_status=payment_status_map.get(row['Payment Status'], 'not_paid'),
                            payment_type=payment_type_map.get(row['Payment Type'], 'cash'),
                            payment_frequency=payment_frequency_map.get(row['Payment Frequency'], 'daily'),
                            in_time=pd.to_datetime(row['In Time']).time() if pd.notna(row['In Time']) else None,
                            out_time=pd.to_datetime(row['Out Time']).time() if pd.notna(row['Out Time']) else None,
                            treatment_1=row.get('Treatment 1', ''),
                            treatment_2=row.get('Treatment 2', ''),
                            treatment_3=row.get('Treatment 3', ''),
                            treatment_4=row.get('Treatment 4', ''),
                            therapist_1=therapist_map.get(row['Therapist 1'], 'dr_basidh'),
                            therapist_2=therapist_map.get(row.get('Therapist 2', ''), None) if pd.notna(row.get('Therapist 2')) else None
                        )
                        sheet.save()
                        imported_count += 1
                    except Exception as e:
                        messages.warning(request, f'Error importing row {index + 2}: {str(e)}')
                        continue
                
                messages.success(request, f'Successfully imported {imported_count} records!')
                return redirect('daily_sheet_list')
                
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'daily_sheet_import.html', {'form': form})
